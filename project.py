## IMPORTING LIBRARIES
import win32com.client
import pandas as pd
import re
from openpyxl.styles import PatternFill

# ------------------------------------
# CONNECT TO OUTLOOK
# ------------------------------------

outlook = win32com.client.Dispatch("Outlook.Application")
namespace = outlook.GetNamespace("MAPI")

inbox = namespace.GetDefaultFolder(6)
messages = inbox.Items

# ------------------------------------
# GROUP EMAILS BY CONVERSATION
# ------------------------------------

threads = {}

for message in messages:

    try:

        convo_id = str(message.ConversationID)

        sender = ""

        try:
            sender = str(message.SenderEmailAddress).lower()
        except:
            sender = ""

        try:
            received_time = message.ReceivedTime
        except:
            received_time = None

        email_data = {
            "sender": sender,
            "subject": str(message.Subject),
            "body": str(message.Body),
            "date": received_time
        }

        if convo_id not in threads:
            threads[convo_id] = []

        threads[convo_id].append(email_data)

    except Exception as e:
        continue

# ------------------------------------
# HELPER FUNCTION
# ------------------------------------

def extract(pattern, text):

    try:
        match = re.search(pattern, text, re.IGNORECASE)

        if match:
            return match.group(1).strip()

        return ""

    except:
        return ""

# ------------------------------------
# PROCESS ORDERS
# ------------------------------------

rows = []

for convo_id, emails in threads.items():

    try:

        # remove mails without dates
        emails = [e for e in emails if e["date"] is not None]

        if len(emails) == 0:
            continue

        # oldest -> newest
        emails.sort(key=lambda x: x["date"])

        first_mail = emails[0]

        # ONLY process if first mail is from ecomm
        if first_mail["sender"] != "ecomm1@ritvaa.in":
            continue

        body = first_mail["body"]

        # --------------------------------
        # ORDER DETAILS
        # --------------------------------

        order_id = extract(
            r"Order Number\s*:\s*(.+)",
            body
        )


        customer_name = extract(
            r"Customer Name\s*:\s*(.+)",
            body
        )

        product_name = extract(
            r"Product Name\s*:\s*(.+)",
            body
        )

        sku_code = extract(
            r"SKU Code\s*:\s*(.+)",
            body
        )

        quantity = extract(
            r"Quantity\s*:\s*(.+)",
            body
        )

        contact_no = extract(
            r"Customer Contact No\.?\s*:\s*(.+)",
            body
        )

        order_date = extract(
            r"Order Date\s*:\s*(.+)",
            body
        )

        # --------------------------------
        # VALIDATE ORDER MAIL
        # --------------------------------

        filled_fields = 0

        if order_id:
            filled_fields += 1

        if customer_name:
            filled_fields += 1

        if product_name:
            filled_fields += 1

        if sku_code:
            filled_fields += 1

        if quantity:
            filled_fields += 1

        # Ignore non-order mails
        if filled_fields < 3:
            continue

        # --------------------------------
        # SUBJECT + SHOPIFY ID
        # --------------------------------

        subjectline = first_mail["subject"]

        shopify_id = ""

        match = re.search(
            r"M\d+",
            subjectline
        )

        if match:
            shopify_id = match.group()

        # --------------------------------
        # MAIL DATES
        # --------------------------------

        mail_date = first_mail["date"].strftime("%Y-%m-%d")

        crm_mail = "NO"
        crm_date = ""

        inventory_mail = "NO"
        inventory_date = ""

        initial_crm_status = ""
        final_crm_status = ""

        # --------------------------------
        # REPLY CHECK
        # --------------------------------

        crm_status = "PENDING"

        confirm_keywords = [
            "order confirmed",
            "customer confirmed",
            "proceed with qc",
            "proceed with the qc",
            "proceed with dispatch",
            "qc and dispatch",
            "kindly proceed"
        ]

        cancel_keywords = [
            "cancel",
            "cancelled",
            "canceled",
            "cancellation"
        ]

        replacement_keywords = [
            "replacement",
            "replace",
            "replaced"
        ]

        for mail in emails:

            sender = mail["sender"]

            text = (
                str(mail["subject"]) + " " +
                str(mail["body"])
            ).lower()

            # CRM MAIL
            if sender == "crmteam@ritvaa.in":

                if any(word in text for word in cancel_keywords):
                    
                    crm_mail = "YES"
                    detected_status = "CANCELLED"

                    if initial_crm_status == "":
                        initial_crm_status = detected_status

                    final_crm_status = detected_status

                    if crm_date == "":
                        crm_date = mail["date"].strftime("%Y-%m-%d")

                elif any(word in text for word in replacement_keywords):

                    crm_mail = "YES"
                    detected_status = "REPLACEMENT"

                    if initial_crm_status == "":
                        initial_crm_status = detected_status

                    final_crm_status = detected_status

                    if crm_date == "":
                        crm_date = mail["date"].strftime("%Y-%m-%d")

                elif any(word in text for word in confirm_keywords):

                    crm_mail = "YES"
                    detected_status = "CONFIRMED"

                    if initial_crm_status == "":
                        initial_crm_status = detected_status

                    final_crm_status = detected_status

                    if crm_date == "":
                        crm_date = mail["date"].strftime("%Y-%m-%d")

            # INVENTORY MAIL
            elif sender == "inventory2@ritvaa.in":

                inventory_mail = "YES"

                if inventory_date == "":
                    inventory_date = mail["date"].strftime("%Y-%m-%d")

        # --------------------------------
        # STATUS
        # --------------------------------

        if final_crm_status == "CANCELLED":

            status = "CANCELLED"

        elif final_crm_status == "REPLACEMENT":

            status = "REPLACEMENT"

        elif inventory_mail == "YES":

            status = "DELIVERED"

        elif initial_crm_status == "CONFIRMED":

            status = "ORDER CONFIRMED"

        else:

            status = "ORDER PENDING"

        # --------------------------------
        # FINAL ROW
        # --------------------------------

        rows.append({
            "ConversationID": convo_id,
            "Order_ID": order_id,
            "Shopify_ID": shopify_id,
            "Customer_Name": customer_name,
            "Product_Name": product_name,
            "SKU_Code": sku_code,
            "Quantity": quantity,
            "Customer_Contact_No": contact_no,
            "Order_Date": order_date,
            "Mail_Date": mail_date,
            "CRM_Mail": crm_mail,
            "CRM_Date": crm_date,
            "Inventory_Mail": inventory_mail,
            "Inventory_Date": inventory_date,
            "Status": status,
            "SubjectLine": subjectline
        })

    except Exception as e:
        print("Error processing conversation:", convo_id)
        print(e)

# ------------------------------------
# EXPORT TO EXCEL
# ------------------------------------

print("Orders Found:", len(rows))

df = pd.DataFrame(rows)

# ------------------------------------
# SUMMARY SHEET DATA
# ------------------------------------

from datetime import datetime, timedelta

# ------------------------------------
# DATE CONVERSIONS
# ------------------------------------

df["Mail_Date_Raw"] = pd.to_datetime(
    df["Mail_Date"],
    dayfirst=True,
    errors="coerce"
)

today = datetime.today().date()
yesterday = today - timedelta(days=1)

# Current week
current_week = today.isocalendar()[1]
current_year = today.year

# Last week
last_week_date = today - timedelta(days=7)
last_week = last_week_date.isocalendar()[1]

# Current month
current_month = today.month

# Last month
if current_month == 1:
    last_month = 12
    last_month_year = current_year - 1
else:
    last_month = current_month - 1
    last_month_year = current_year

# ------------------------------------
# FILTERS
# ------------------------------------

today_df = df[
    df["Mail_Date_Raw"].dt.date == today
]

yesterday_df = df[
    df["Mail_Date_Raw"].dt.date == yesterday
]

this_week_df = df[
    (df["Mail_Date_Raw"].dt.isocalendar().week == current_week)
]

last_week_df = df[
    (df["Mail_Date_Raw"].dt.isocalendar().week == last_week)
]

this_month_df = df[
    (df["Mail_Date_Raw"].dt.month == current_month) &
    (df["Mail_Date_Raw"].dt.year == current_year)
]

last_month_df = df[
    (df["Mail_Date_Raw"].dt.month == last_month) &
    (df["Mail_Date_Raw"].dt.year == last_month_year)
]

# ------------------------------------
# SUMMARY FUNCTION
# ------------------------------------

def get_summary(dataframe):

    return [
        len(dataframe),
        len(dataframe[dataframe["Status"] == "ORDER PENDING"]),
        len(dataframe[dataframe["Status"] == "ORDER CONFIRMED"]),
        len(dataframe[dataframe["Status"] == "DELIVERED"]),
        len(dataframe[dataframe["Status"] == "CANCELLED"]),
        len(dataframe[dataframe["Status"] == "REPLACEMENT"])
    ]
# ------------------------------------
# TABLE 1
# ------------------------------------

today_vs_yesterday = pd.DataFrame({
    "Metric": [
        "No. of Orders",
        "Pending",
        "Confirmed",
        "Delivered",
        "Cancelled",
        "Replacement"
    ],
    "Today": get_summary(today_df),
    "Yesterday": get_summary(yesterday_df)
})

# ------------------------------------
# TABLE 2
# ------------------------------------

week_vs_week = pd.DataFrame({
    "Metric": [
        "No. of Orders",
        "Pending",
        "Confirmed",
        "Delivered",
        "Cancelled",
        "Replacement"
    ],
    "This Week": get_summary(this_week_df),
    "Last Week": get_summary(last_week_df)
})

# ------------------------------------
# TABLE 3
# ------------------------------------

month_vs_month = pd.DataFrame({
    "Metric": [
        "No. of Orders",
        "Pending",
        "Confirmed",
        "Delivered",
        "Cancelled",
        "Replacement"
    ],
    "This Month": get_summary(this_month_df),
    "Last Month": get_summary(last_month_df)
})

#aaaaaaaa

df = df.fillna("")

for col in df.columns:
    df[col] = df[col].astype(str)

print(df.head())

import os

if os.path.exists("Order_Tracking.xlsx"):
    os.remove("Order_Tracking.xlsx")


# ====================================
# WEEKLY OPERATIONS DASHBOARD
# ====================================

week_start = today - timedelta(days=today.weekday())

week_dates = [
    week_start + timedelta(days=i)
    for i in range(7)
]

date_headers = [
    d.strftime("%d-%b")
    for d in week_dates
]

day_headers = [
    d.strftime("%A")
    for d in week_dates
]

weekly_dashboard = {
    "Metric": [
        "Day",
        "Orders",
        "Pending",
        "Confirmed",
        "Delivered",
        "Cancelled",
        "Replacement"
    ]
}

for current_day in week_dates:

    orders = len(
        df[
            pd.to_datetime(
                df["Mail_Date"],
                errors="coerce"
            ).dt.date == current_day
        ]
    )

    pending = len(
        df[
            df["Status"] == "ORDER PENDING"
        ]
    )

    confirmed = len(
        df[
            pd.to_datetime(
                df["CRM_Date"],
                errors="coerce"
            ).dt.date == current_day
        ]
    )

    delivered = len(
        df[
            pd.to_datetime(
                df["Inventory_Date"],
                errors="coerce"
            ).dt.date == current_day
        ]
    )

    cancelled = len(
        df[
            (df["Status"] == "CANCELLED") &
            (
                pd.to_datetime(
                    df["CRM_Date"],
                    errors="coerce"
                ).dt.date == current_day
            )
        ]
    )

    replacement = len(
        df[
            (df["Status"] == "REPLACEMENT") &
            (
                pd.to_datetime(
                    df["CRM_Date"],
                    errors="coerce"
                ).dt.date == current_day
            )
        ]
    )

    weekly_dashboard[
        current_day.strftime("%d-%b")
    ] = [
        current_day.strftime("%A"),
        orders,
        pending,
        confirmed,
        delivered,
        cancelled,
        replacement
    ]

weekly_dashboard_df = pd.DataFrame(
    weekly_dashboard
)

from openpyxl.styles import Font, Border, Side

with pd.ExcelWriter(
    "Order_Tracking.xlsx",
    engine="openpyxl"
) as writer:

    weekly_dashboard_df.to_excel(
        writer,
        sheet_name="Weekly_Report",
        index=False
    )

    weekly_ws = writer.sheets["Weekly_Report"]

    # Sheet 1
    today_vs_yesterday.to_excel(
        writer,
        sheet_name="Summary",
        startrow=1,
        startcol=0,
        index=False
    )

    week_vs_week.to_excel(
        writer,
        sheet_name="Summary",
        startrow=1,
        startcol=5,
        index=False
    )

    month_vs_month.to_excel(
        writer,
        sheet_name="Summary",
        startrow=1,
        startcol=10,
        index=False
    )

    # Sheet 2
    df.to_excel(
        writer,
        sheet_name="Order_Data",
        index=False
    )

    today_df.to_excel(
        writer,
        sheet_name="Today_Orders",
        index=False
    )

    yesterday_df.to_excel(
        writer,
        sheet_name="Yesterday_Orders",
        index=False
    )

    this_week_df.to_excel(
        writer,
        sheet_name="This_Week_Orders",
        index=False
    )

    last_week_df.to_excel(
        writer,
        sheet_name="Last_Week_Orders",
        index=False
    )

    this_month_df.to_excel(
        writer,
        sheet_name="This_Month_Orders",
        index=False
    )

    last_month_df.to_excel(
        writer,
        sheet_name="Last_Month_Orders",
        index=False
    )

    weekly_dashboard_df.to_excel(
        writer,
        sheet_name="Weekly_Report",
        index=False
    )

    # ==========================
    # FONT FORMATTING
    # ==========================

    summary_ws = writer.sheets["Summary"]
    # Section Titles

    summary_ws["A1"] = "TODAY vs YESTERDAY"

    summary_ws["F1"] = "THIS WEEK vs LAST WEEK"

    
    summary_ws["K1"] = "THIS MONTH vs LAST MONTH"
    order_ws = writer.sheets["Order_Data"]

    today_ws = writer.sheets["Today_Orders"]
    yesterday_ws = writer.sheets["Yesterday_Orders"]
    this_week_ws = writer.sheets["This_Week_Orders"]
    last_week_ws = writer.sheets["Last_Week_Orders"]
    this_month_ws = writer.sheets["This_Month_Orders"]
    last_month_ws = writer.sheets["Last_Month_Orders"]

    detail_sheets = [
        order_ws,
        today_ws,
        yesterday_ws,
        this_week_ws,
        last_week_ws,
        this_month_ws,
        last_month_ws,
        weekly_ws
    ]
    # TEST HYPERLINK

    summary_ws["B2"].hyperlink = "#'Today_Orders'!A1"
    summary_ws["B2"].style = "Hyperlink"

        # Today / Yesterday
    
    summary_ws["C2"].hyperlink = "#Yesterday_Orders!A1"
    summary_ws["C2"].style = "Hyperlink"

    # This Week / Last Week
    summary_ws["G2"].hyperlink = "#This_Week_Orders!A1"
    summary_ws["G2"].style = "Hyperlink"
    summary_ws["H2"].hyperlink = "#Last_Week_Orders!A1"
    summary_ws["H2"].style = "Hyperlink"

    # This Month / Last Month
    summary_ws["L2"].hyperlink = "#This_Month_Orders!A1"
    summary_ws["L2"].style = "Hyperlink"
    summary_ws["M2"].hyperlink = "#Last_Month_Orders!A1"
    summary_ws["M2"].style = "Hyperlink"

    # Summary Sheet
    for row in summary_ws.iter_rows():
        for cell in row:
            cell.font = Font(
                name="Century Gothic",
                size=10
            )

    # Order Data Sheet
    for ws in detail_sheets:

        for row in ws.iter_rows():

            for cell in row:

                cell.font = Font(
                    name="Century Gothic",
                    size=10
                )

    # ==========================
    # HEADER FORMATTING
    # ==========================

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78"
    )
    for ws in detail_sheets:
        for cell in ws[1]:

            cell.fill = header_fill

            cell.font = Font(
                name="Century Gothic",
                size=10,
                bold=True,
                color="FFFFFF"
            )

    # Border Style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Apply Border to Summary Sheet
    for row in summary_ws.iter_rows():
        for cell in row:
            cell.border = thin_border


    # Apply Border to Order Data Sheet
    for ws in detail_sheets:

        for row in ws.iter_rows():

            for cell in row:

                cell.border = thin_border

        weekly_ws = writer.sheets["Weekly_Report"]

    for row in weekly_ws.iter_rows():

        for cell in row:

            cell.font = Font(
                name="Century Gothic",
                size=10
            )

    for cell in weekly_ws[1]:

        cell.fill = header_fill

        cell.font = Font(
            name="Century Gothic",
            size=10,
            bold=True,
            color="FFFFFF"
        )

    # ==========================
    # SUMMARY TITLE FORMATTING
    # ==========================

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78"
    )

    white_font = Font(
        name="Century Gothic",
        size=10,
        bold=True,
        color="FFFFFF"
    )

    for cell_ref in ["A1", "F1", "K1"]:

        summary_ws[cell_ref].fill = header_fill

        summary_ws[cell_ref].font = white_font

    # ==========================
    # SUMMARY TABLE HEADER ROW
    # ==========================

    for col in ["A", "B", "C"]:

        summary_ws[f"{col}2"].fill = header_fill

        summary_ws[f"{col}2"].font = white_font


    for col in ["F", "G", "H"]:

        summary_ws[f"{col}2"].fill = header_fill

        summary_ws[f"{col}2"].font = white_font


    for col in ["K", "L", "M"]:

        summary_ws[f"{col}2"].fill = header_fill

        summary_ws[f"{col}2"].font = white_font
    
    # ==========================
    # REMOVE GAP COLUMN BORDERS
    # ==========================

    from openpyxl.styles import Border

    no_border = Border()

    for col in ["D", "E", "I", "J"]:

        for row in range(
            1,
            summary_ws.max_row + 1
        ):

            summary_ws[
                f"{col}{row}"
            ].border = no_border
    # ==========================
    # STATUS COLOR FORMATTING
    # ==========================

    from openpyxl.styles import PatternFill

    # ==========================
    # STATUS COLOR FORMATTING
    # ==========================

    green_fill = PatternFill(
        fill_type="solid",
        start_color="C6EFCE"
    )

    yellow_fill = PatternFill(
        fill_type="solid",
        start_color="FFEB9C"
    )

    red_fill = PatternFill(
        fill_type="solid",
        start_color="FFC7CE"
    )

    for ws in detail_sheets:

        status_col = None

        # Find Status Column
        for cell in ws[1]:

            if str(cell.value).strip() == "Status":

                status_col = cell.column

                break

        # Skip sheet if Status column not found
        if status_col is None:
            continue

        # Color Status Cells
        for row in range(2, ws.max_row + 1):

            status_cell = ws.cell(
                row=row,
                column=status_col
            )

            if status_cell.value == "DELIVERED":

                status_cell.fill = green_fill

            elif status_cell.value == "ORDER CONFIRMED":

                status_cell.fill = yellow_fill

            elif status_cell.value == "ORDER PENDING":

                status_cell.fill = red_fill

            elif status_cell.value == "CANCELLED":

                status_cell.fill = red_fill

            elif status_cell.value == "REPLACEMENT":

                status_cell.fill = yellow_fill

    # ==========================
    # AUTO COLUMN WIDTH
    # ==========================

    for ws in [summary_ws] + detail_sheets:

        for column in ws.columns:

            max_length = 0

            for cell in column:

                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = max_length + 2

            ws.column_dimensions[
                column[0].column_letter
            ].width = adjusted_width